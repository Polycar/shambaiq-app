import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_beautiful_excel_month2():
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Define styles
    font_family = "Segoe UI"
    
    # Fonts
    title_font = Font(name=font_family, size=16, bold=True, color="1E3A1E")
    section_font = Font(name=font_family, size=12, bold=True, color="1E3A1E")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    bold_body_font = Font(name=font_family, size=10, bold=True, color="333333")
    body_font = Font(name=font_family, size=10, color="333333")
    italic_body_font = Font(name=font_family, size=10, italic=True, color="666666")
    link_font = Font(name=font_family, size=10, underline="single", color="1E3A1E")
    
    # Fills
    header_fill = PatternFill(start_color="1E3A1E", end_color="1E3A1E", fill_type="solid")
    alt_row_fill = PatternFill(start_color="F1F6F1", end_color="F1F6F1", fill_type="solid")
    accent_fill = PatternFill(start_color="E8F2E8", end_color="E8F2E8", fill_type="solid")
    
    # Borders
    thin_border_side = Side(border_style="thin", color="E0E0E0")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    thick_bottom_side = Side(border_style="medium", color="1E3A1E")
    header_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thick_bottom_side)
    
    # Alignments
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    title_align = Alignment(horizontal="left", vertical="center")

    # ==========================================
    # SHEET 1: STRATEGY BLUEPRINT (MONTH 2)
    # ==========================================
    ws1 = wb.active
    ws1.title = "Strategy Blueprint Month 2"
    ws1.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws1.cell(row=2, column=2, value="ShambaIQ SEO Content Strategy - Month 2 (Days 31-60)").font = title_font
    ws1.cell(row=2, column=2).alignment = title_align
    ws1.row_dimensions[2].height = 30
    
    # Strategy Section
    ws1.cell(row=4, column=2, value="Core Strategy Blueprint").font = section_font
    ws1.cell(row=5, column=2, value="This second month focuses on high-converting niche horticultural, cash, fodder, and new specialty crops across various Kenyan agroecological zones, directing targeted web traffic directly to pre-configured inputs and soil intelligence views.").font = body_font
    ws1.cell(row=5, column=2).alignment = left_align
    ws1.row_dimensions[5].height = 40
    
    # Objective & KPI Table
    ws1.cell(row=7, column=2, value="SEO Key Performance Indicators (KPIs)").font = section_font
    
    headers1 = ["Objective Category", "Detailed Month 2 KPI", "Target Action/Funnels"]
    for col_idx, h in enumerate(headers1, start=2):
        cell = ws1.cell(row=8, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = header_border
    ws1.row_dimensions[8].height = 25
    
    kpis = [
        ("Niche Crops Focus", "Establish domain authority for regional specialty value chains including Lowland Paddy Rice, Watermelon, Green Grams, Cashew Nuts, Capsicum, Fodders, and Orchard crops.", "Targeted guides using the newly synchronized crops in ShambaIQ"),
        ("Dynamic Conversion Funnel", "Direct growers from regional search landers directly into specific input credit and soil diagnostics tools with pre-filled query parameters.", "Embedded `/app?county=X&crop=Y` links customized to specialty crops"),
        ("EEAT Quality Assurance", "Integrate actual regional soil baseline diagnostics, micro-nutrient requirements, and certified seed references into every post.", "County-specific soil profiles and KALRO seed varieties validation")
    ]
    
    for row_idx, kpi in enumerate(kpis, start=9):
        fill = alt_row_fill if row_idx % 2 == 0 else PatternFill(fill_type=None)
        for col_idx, val in enumerate(kpi, start=2):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.font = bold_body_font if col_idx == 2 else body_font
            if fill.fill_type:
                cell.fill = fill
            cell.alignment = left_align
            cell.border = thin_border
        ws1.row_dimensions[row_idx].height = 40
        
    ws1.column_dimensions['A'].width = 3
    ws1.column_dimensions['B'].width = 25
    ws1.column_dimensions['C'].width = 65
    ws1.column_dimensions['D'].width = 45

    # ==========================================
    # SHEET 2: MONTH 2 CONTENT CALENDAR
    # ==========================================
    ws2 = wb.create_sheet(title="Month 2 Content Calendar")
    ws2.views.sheetView[0].showGridLines = True
    
    ws2.cell(row=2, column=2, value="Month 2 Content Calendar: Days 31-60").font = title_font
    ws2.row_dimensions[2].height = 30
    
    headers2 = [
        "Day Range", "Topic Title", "Target Focus Keyword", 
        "Secondary Keywords", "Search Intent", 
        "County Soil Integration & Guidelines", "Custom ShambaIQ CTA Text", "Advisory CTA Link"
    ]
    
    for col_idx, h in enumerate(headers2, start=2):
        cell = ws2.cell(row=4, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = header_border
    ws2.row_dimensions[4].height = 30
    
    calendar_data = [
        (
            "Day 31-33", 
            "Kisumu County Lowland Paddy Rice Optimization: Double Your Harvests", 
            "rice farming in Kisumu", 
            "lowland paddy rice Mwea alternative, Kisumu soil silicon, best fertilizer for paddy rice", 
            "Commercial & Transactional",
            "- Lake Victoria basin heavy silty clay soils with exceptional water retention.\n- Guidelines: Address low soil silicon and nitrogen leaching under saturated conditions. Recommend customized basal NPK followed by split Urea/Ammonium Sulfate applications. Select certified high-yielding varieties.",
            "Optimize Your Paddy: Calculate the exact urea split-application intervals. Use the ShambaIQ Lowland Rice Tool.",
            "http://localhost:3000/app?county=Kisumu&crop=Rice%20(Lowland/Paddy)"
        ),
        (
            "Day 34-36", 
            "High-Yield Watermelon Farming in Garissa County: A Tana Basin Goldmine", 
            "watermelon farming in Garissa", 
            "sandy loam soils watermelon, drip irrigation Garissa KES, potassium rich fertilizer watermelon", 
            "Informational & Commercial",
            "- Arid alluvial sandy loam soils of the Tana River basin.\n- Guidelines: Focus on dryland drip irrigation parameters and low organic carbon levels. Detail phosphorus basal requirements and high-potassium top-dressing rules to maximize fruit size and sugar content.",
            "Get Your Watermelon Budget: Maximize fruit weight and sugar content. Run the ShambaIQ Garissa Watermelon Tool.",
            "http://localhost:3000/app?county=Garissa&crop=Watermelon"
        ),
        (
            "Day 37-39", 
            "Double Harvest Green Grams (Ndengu) in Kitui: High-Yield Dryland Guide", 
            "green grams farming in Kitui", 
            "drought resistant legumes Kitui, ndengu spacing Kenya, phosphorus booster Ndengu", 
            "Informational & Sustainable",
            "- Light sandy loam semi-arid alfisols with low organic carbon and low baseline P.\n- Guidelines: Emphasize seed inoculation with Rhizobium to harness biological nitrogen fixation. Prescribe low-cost SSP/phosphate basal inputs, avoiding raw nitrogen feeds to protect root nodules.",
            "Cut Input Costs: Let green grams fix nitrogen for you. Calculate optimal seed-dressing and SSP rates with the Kitui Ndengu Tool.",
            "http://localhost:3000/app?county=Kitui&crop=Green%20Grams"
        ),
        (
            "Day 40-42", 
            "Coastal Cashew Nut Production in Kilifi County: Revitalizing Soils", 
            "cashew nut farming in Kilifi", 
            "coastal soils cashew tree, cashew nut yield per acre Kenya, NPK schedule cashew", 
            "Commercial & Transactional",
            "- Highly leached, low-CEC coastal sandy soils.\n- Guidelines: Explain how high sand content triggers rapid leaching. Recommend split applications of organic manure compost alongside customized slow-release compound fertilizers. Select high-yielding certified coastal clones.",
            "Maximize Orchard Output: Ensure nutrients do not leach away. Plan your coastal cashew tree nutrition with ShambaIQ.",
            "http://localhost:3000/app?county=Kilifi&crop=Cashew%20Nuts"
        ),
        (
            "Day 43-45", 
            "Greenhouse Capsicum Production in Kajiado: Maximizing Quality & Sizes", 
            "capsicum farming in Kajiado", 
            "greenhouse capsicum budget Kenya, soil pH capsicum, foliar feed calcium capsicum", 
            "Informational & Commercial",
            "- Alkaline clay soils with low available zinc and iron.\n- Guidelines: Highlight how alkaline pH blocks trace mineral uptake. Prescribe ammonium sulfate basal application to buffer pH downwards and select calcium-rich foliar sprays to eliminate blossom end rot.",
            "Optimize Greenhouse Feeds: Maximize crop sizes and prevent blossom end rot. Appraise your farm at ShambaIQ Kajiado Capsicum Advisor.",
            "http://localhost:3000/app?county=Kajiado&crop=Capsicum"
        ),
        (
            "Day 46-48", 
            "Avocado & Fodder Integrated Orchards in Murang'a County", 
            "avocado farming in Muranga", 
            "avocado orchard spacing Kenya, Muranga red volcanic soils, organic manure avocado", 
            "Informational & Sustainable",
            "- Deep red volcanic clay loams rich in organic matter but moderately acidic.\n- Guidelines: Address deep taproot biological mining. Emphasize organic composting schedules, basal phosphate application to stimulate root anchors, and crop integration with legume fodder.",
            "Maximize Avocado Output: Optimize long-term volcanic soil inputs. Try the ShambaIQ Murang'a Avocado Tool.",
            "http://localhost:3000/app?county=Murang-a&crop=Avocado"
        ),
        (
            "Day 49-51", 
            "High-Yield Sugarcane Farming in Bungoma: Soil Restoration Guide", 
            "sugarcane farming in Bungoma", 
            "Bungoma soil acidity sugarcane, fertilizer rates sugarcane Nzoia, high nitrogen topdress sugarcane", 
            "Commercial & Transactional",
            "- Moderately acidic weathered loam soils with high cumulative nitrogen depletion.\n- Guidelines: Focus on high-tonnage nitrogen top-dressing schedules. Detail mechanical tillage practices, trash-incorporation to restore soil organic matter, and dolomitic lime buffering.",
            "Double Your Sugarcane Tonnage: Protect your crop from soil acidification. Plan Bungoma sugarcane inputs with ShambaIQ.",
            "http://localhost:3000/app?county=Bungoma&crop=Sugarcane"
        ),
        (
            "Day 52-54", 
            "Lucerne Fodder Optimization in Nakuru: Smart Dairy Integration", 
            "lucerne farming in Nakuru", 
            "best soil for lucerne Kenya, lucerne seed rate per acre, phosphate fertilizer lucerne", 
            "Informational & Commercial",
            "- Well-drained volcanic ash loam soils with neutral to slightly alkaline pH.\n- Guidelines: Detail lucerne's massive phosphorus demand for nodule development. Prescribe heavy basal phosphate inputs and highlight crop rotations to feed nitrogen directly to subsequent silage maize crops.",
            "Double Homegrown Protein: Build a multi-year high-yield fodder bank. Calculate your lucerne input needs with Nakuru Lucerne Optimizer.",
            "http://localhost:3000/app?county=Nakuru&crop=Lucerne"
        ),
        (
            "Day 55-57", 
            "Pixie Oranges Dryland Cultivation in Makueni: Soil & Crop Optimization", 
            "pixie oranges farming in Makueni", 
            "pixie oranges yield per acre, Makueni soil pH citrus, micro-nutrient foliar spray orange", 
            "Informational & Transactional",
            "- Semi-arid sandy clay loams, low in available organic carbon and vital micronutrients.\n- Guidelines: Highlight K-rich formulations to support fruit cell walls. Recommend specific zinc and boron foliar inputs at early flowering to double fruit set and prevent premature fruit drop.",
            "Boost Citrus Profits: Protect dryland orange trees from micronutrient deficits. Appraise your farm at ShambaIQ Makueni Pixie Advisor.",
            "http://localhost:3000/app?county=Makueni&crop=Pixie%20Oranges"
        ),
        (
            "Day 58-60", 
            "Wambugu Apples Highland Cultivation in Nyeri: Organic & Chemical Success", 
            "apple farming in Nyeri", 
            "wambugu apple soil requirements, chill hours apples Kenya, organic mulching apple orchards", 
            "Informational",
            "- Cool highland volcanic clay loams with high baseline organic carbon.\n- Guidelines: Focus on cold-air drainage zones and rich organic mulching to protect roots. Prescribe exact high-phosphate basal inputs at tree planting and organic compost schedules to ensure strong wood structure.",
            "Plan Your Apple Orchard: Establish healthy, multi-decade highland apple stands. Calculate your inputs with the Nyeri Apple Advisor.",
            "http://localhost:3000/app?county=Nyeri&crop=Wambugu%20Apples"
        )
    ]
    
    for row_idx, row_data in enumerate(calendar_data, start=5):
        fill = alt_row_fill if row_idx % 2 == 0 else PatternFill(fill_type=None)
        
        # Day
        cell_day = ws2.cell(row=row_idx, column=2, value=row_data[0])
        cell_day.font = bold_body_font
        cell_day.alignment = center_align
        cell_day.border = thin_border
        if fill.fill_type:
            cell_day.fill = fill
            
        # Title
        cell_title = ws2.cell(row=row_idx, column=3, value=row_data[1])
        cell_title.font = bold_body_font
        cell_title.alignment = left_align
        cell_title.border = thin_border
        if fill.fill_type:
            cell_title.fill = fill
            
        # Focus Keyword
        cell_kw = ws2.cell(row=row_idx, column=4, value=row_data[2])
        cell_kw.font = italic_body_font
        cell_kw.alignment = left_align
        cell_kw.border = thin_border
        if fill.fill_type:
            cell_kw.fill = fill
            
        # Secondary Keywords
        cell_sec = ws2.cell(row=row_idx, column=5, value=row_data[3])
        cell_sec.font = body_font
        cell_sec.alignment = left_align
        cell_sec.border = thin_border
        if fill.fill_type:
            cell_sec.fill = fill
            
        # Search Intent
        cell_intent = ws2.cell(row=row_idx, column=6, value=row_data[4])
        cell_intent.font = body_font
        cell_intent.alignment = center_align
        cell_intent.border = thin_border
        if fill.fill_type:
            cell_intent.fill = fill
            
        # Soil Integration
        cell_soil = ws2.cell(row=row_idx, column=7, value=row_data[5])
        cell_soil.font = body_font
        cell_soil.alignment = left_align
        cell_soil.border = thin_border
        if fill.fill_type:
            cell_soil.fill = fill
            
        # CTA Text
        cell_cta = ws2.cell(row=row_idx, column=8, value=row_data[6])
        cell_cta.font = body_font
        cell_cta.alignment = left_align
        cell_cta.border = thin_border
        if fill.fill_type:
            cell_cta.fill = fill
            
        # CTA Link
        cell_link = ws2.cell(row=row_idx, column=9, value="Open ShambaIQ Advisory Tool")
        cell_link.font = link_font
        cell_link.alignment = center_align
        cell_link.hyperlink = row_data[7]
        cell_link.border = thin_border
        if fill.fill_type:
            cell_link.fill = fill
            
        # Calculate row height
        max_len = max(len(row_data[1]), len(row_data[5]), len(row_data[6]))
        ws2.row_dimensions[row_idx].height = max(45, max_len // 3)

    # Column widths
    ws2.column_dimensions['A'].width = 3
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 32
    ws2.column_dimensions['D'].width = 24
    ws2.column_dimensions['E'].width = 32
    ws2.column_dimensions['F'].width = 20
    ws2.column_dimensions['G'].width = 50
    ws2.column_dimensions['H'].width = 45
    ws2.column_dimensions['I'].width = 25

    # ==========================================
    # SHEET 3: MONTH 2 SEO CHECKLIST
    # ==========================================
    ws3 = wb.create_sheet(title="SEO Checklist")
    ws3.views.sheetView[0].showGridLines = True
    
    ws3.cell(row=2, column=2, value="SEO Execution Checklist for Content Creators (Month 2)").font = title_font
    ws3.row_dimensions[2].height = 30
    
    headers3 = ["Audit Check", "Detailed Guideline & Action Step", "Verified (Y/N)"]
    for col_idx, h in enumerate(headers3, start=2):
        cell = ws3.cell(row=4, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = header_border
    ws3.row_dimensions[4].height = 25
    
    checklist = [
        ("1. Niche Keywords Integration", "Ensure the primary focus keyword (e.g. 'watermelon farming in Garissa') appears at the start of the title, first H2, and post excerpt.", ""),
        ("2. Real Soil Values Verification", "Verify the soil pH and nutrient levels in the post match the county values in src/lib/site-data.ts and the backend database.", ""),
        ("3. Precision Advisory Dynamic Links", "Every dynamic CTA link must contain the matching parameters to pre-fill the tool accurately (e.g., /app?county=Garissa&crop=Watermelon).", ""),
        ("4. Complementary Root Interlinks", "Link to other relevant specialty crop guides or county soil profiles (e.g., green grams in Kitui linking to Machakos soil restoration).", "")
    ]
    
    for row_idx, item in enumerate(checklist, start=5):
        fill = alt_row_fill if row_idx % 2 == 0 else PatternFill(fill_type=None)
        for col_idx, val in enumerate(item, start=2):
            cell = ws3.cell(row=row_idx, column=col_idx, value=val)
            cell.font = bold_body_font if col_idx == 2 else body_font
            if fill.fill_type:
                cell.fill = fill
            cell.alignment = center_align if col_idx == 4 else left_align
            cell.border = thin_border
        ws3.row_dimensions[row_idx].height = 40
        
    ws3.column_dimensions['A'].width = 3
    ws3.column_dimensions['B'].width = 30
    ws3.column_dimensions['C'].width = 75
    ws3.column_dimensions['D'].width = 15

    # Save to file
    filename = "d:\\shambaiq-app\\seo_content_calendar_month2.xlsx"
    wb.save(filename)
    print(f"Beautifully styled Excel calendar for Month 2 saved successfully to: {filename}")

if __name__ == "__main__":
    create_beautiful_excel_month2()
